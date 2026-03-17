from tkinter import *
from tkinter.font import Font
from tkinter import filedialog
from tkinter import ttk
import json
import os
import sys
from datetime import datetime, timedelta

win = Tk()
win.title("제주 옵서버스 보고서 작성 자동화 프로그램")
win.geometry("450x400")

EXE_PATH = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.getcwd()
SAVE_FILE = os.path.join(EXE_PATH, "settings.json")

def save_settings(data):
    with open(SAVE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def load_settings():
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

settings = load_settings()

def create_label_and_entry(win, label_name, row, col, default_text=""):
    label = Label(win, text=label_name, font=Font(family="KoPub돋움체 bold", size = 10))
    label.grid(column=col, row=row, padx=20, pady=10)

    entry = Entry(win, width=15, font=Font(family="KoPub돋움체 bold", size = 10))
    entry.grid(column=col + 1, row=row, padx=0, pady=10)
    entry.insert(0, default_text)
    return entry

def create_button_and_entry_select_file(win, label_text, row, col, setting_key, previous_path=""):
    def select_file():
        file = filedialog.askdirectory(title="폴더 선택") if "폴더" in label_text else filedialog.askopenfile(
            title="파일 선택",
            filetype=(("csv file", "*.csv"), ("xlsx file", "*.xlsx"), ("all files", "*.*"))
        )
        if file:
            path = file.name if hasattr(file, "name") else file
            entry.delete(0, END)
            entry.insert(0, path)
            settings[setting_key] = path
            save_settings(settings)

    btn = Button(win, text=label_text, command=select_file, width=12, height=1, anchor="w", font=Font(family="KoPub돋움체 bold", size = 10))
    btn.grid(column=col, row=row, padx=20, pady=5)

    entry = Entry(win, width=40, font=Font(family="KoPub돋움체 light", size = 10))
    entry.grid(column=col + 1, row=row, padx=0, pady=0, columnspan=3)
    entry.insert(0, previous_path)
    return entry

def create_button_and_entry_select_file_주중주말(win, label_text, row, col, setting_key, previous_path=""):
    def select_file():
        file = filedialog.askdirectory(title="폴더 선택") if "폴더" in label_text else filedialog.askopenfile(
            title="파일 선택",
            filetype=(("csv file", "*.csv"), ("xlsx file", "*.xlsx"), ("all files", "*.*"))
        )
        if file:
            path = file.name if hasattr(file, "name") else file
            entry.delete(0, END)
            entry.insert(0, path)
            settings[setting_key] = path
            save_settings(settings)

    btn = Button(win, text=label_text, command=select_file, width=12, height=1, anchor="w", font=Font(family="KoPub돋움체 bold", size = 10))
    btn.grid(column=col, row=row, padx=20, pady=5)

    entry = Entry(win, width=20, font=Font(family="KoPub돋움체 light", size = 10))
    entry.grid(column=col + 1, row=row, padx=0, pady=0, columnspan=2)
    entry.insert(0, previous_path)
    return entry


ent_ID = create_label_and_entry(win, "ID", row=1, col=0, default_text="")
ent_PW = create_label_and_entry(win, "PW", row=1, col=2, default_text="")
weekday = datetime.now().weekday()
last_wednesday = datetime.now() - timedelta(days=weekday - 2 + 7)
last_wednesday_str = last_wednesday.strftime("%Y-%m-%d")
ent_시작일 = create_label_and_entry(win, "기간", row=2, col=0, default_text=last_wednesday_str)
start_date_str = ent_시작일.get()
start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
end_date = start_date + timedelta(days=6)
ent_종료일 = create_label_and_entry(win, "~", row=2, col=2, default_text=end_date.strftime("%Y-%m-%d"))


ent_다운로드폴더 = create_button_and_entry_select_file(win, "다운로드 폴더", row=4, col=0, setting_key="다운로드폴더", previous_path=settings.get("다운로드폴더", ""))
ent_Excel_계산용 = create_button_and_entry_select_file(win, "Excel_계산용", row=5, col=0, setting_key="Excel_계산용", previous_path=settings.get("Excel_계산용", ""))
ent_DB_보관용 = create_button_and_entry_select_file(win, "DB_보관용", row=6, col=0, setting_key="DB_보관용", previous_path=settings.get("DB_보관용", ""))
ent_DB_Excel용 = create_button_and_entry_select_file(win, "DB_Excel용", row=7, col=0, setting_key="DB_Excel용", previous_path=settings.get("DB_Excel용", ""))
ent_DB_주중주말 = create_button_and_entry_select_file_주중주말(win, "DB_주중/주말", row=8, col=0, setting_key="DB_주중주말", previous_path=settings.get("DB_주중주말", ""))
ent_한글_보고서 = create_button_and_entry_select_file(win, "한글_보고서", row=9, col=0, setting_key="한글_보고서", previous_path=settings.get("한글_보고서", ""))

def create_button_and_entry_update_DB(win, label_text, row, col):
    def update_DB():
        ID = ent_ID.get()
        PASSWORD = ent_PW.get()
        FROM_DATE = ent_시작일.get()
        TO_DATE = ent_종료일.get()
        PATH_FOLDER_DOWNLOAD = ent_다운로드폴더.get()
        PATH_RAWDATA = ent_DB_보관용.get()
        PATH_EXCEL = ent_Excel_계산용.get()
        PATH_DB_FOR_EXCEL = ent_DB_Excel용.get()
        PATH_HOLIDAY = ent_DB_주중주말.get()
        HOLIDAY = ['2025-01-01']
        import pandas as pd
        def update_weekday_to_weekend(csv_file, target_dates):
            df = pd.read_csv(csv_file, encoding='euc-kr')
            df['날짜'] = pd.to_datetime(df['날짜'])
            for target_date in target_dates:
                mask = (df['날짜'] == pd.to_datetime(target_date)) & (df['주중/주말'] == '주중')
                df.loc[mask, '주중/주말'] = '주말'
            df.drop_duplicates(keep = 'first')
            df.to_csv(csv_file, index=False, encoding='euc-kr')

        update_weekday_to_weekend(PATH_HOLIDAY, HOLIDAY)

        from selenium import webdriver

        driver = webdriver.Chrome()

        driver.get('https://jeju.varo-drt.com/drt3Web/html/login.html')

        from selenium.webdriver.common.by import By
        elem = driver.find_element(By.XPATH, '//*[@id="user_name"]').send_keys(ID)
        elem = driver.find_element(By.XPATH, '//*[@id="user_pwd"]').send_keys(PASSWORD)
        elem = driver.find_element(By.XPATH, '//*[@id="skip-content"]/section/div/form/div[2]/input').click()

        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        wait = WebDriverWait(driver, 10)
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="govSelectName"]')))

        driver.get(f'https://jeju.varo-drt.com/drt3Web/html/run_historylist.html?govSelectPk=1')

        wait = WebDriverWait(driver, 10)
        wait.until(EC.visibility_of_element_located((By.XPATH, '//*[@id="skip-content"]/div/section/form/div[2]/table/tbody/tr')))
        elem = driver.find_element(By.XPATH, '//*[@id="datepicker-s"]')
        elem.clear()
        elem.send_keys(FROM_DATE)
        elem = driver.find_element(By.XPATH, '//*[@id="datepicker-e"]')
        elem.clear()
        elem.send_keys(TO_DATE)

        driver.find_element(By.XPATH, '//*[@id="skip-content"]/div/section/div[1]/form/div[2]/div[2]/a').click()
        import pandas as pd
        df_rawdata = pd.DataFrame(pd.read_csv(PATH_RAWDATA))
        import time
        for PAGE_NUM in range(1, 7):
            if PAGE_NUM == 1:        
                wait = WebDriverWait(driver, 10)
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="skip-content"]/div/section/form/div[2]/table/tbody/tr[1]/td[10]/a')))
                time.sleep(0.5)
                driver.find_element(By.XPATH, '//*[@id="btnExcelDownload"]').click()
            else:
                driver.get(f'https://jeju.varo-drt.com/drt3Web/html/run_historylist.html?govSelectPk={PAGE_NUM}')
                wait = WebDriverWait(driver, 10)
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="skip-content"]/div/section/form/div[2]/table/tbody/tr[1]/td[10]/a')))
                time.sleep(0.5)
                driver.find_element(By.XPATH, '//*[@id="btnExcelDownload"]').click()
            import time
            import pyautogui as pag
            time.sleep(0.5)
            pag.press('enter')
            time.sleep(0.5)
            pag.press('enter')
            import os
            list_name_ctime = []
            for name_file_download in os.listdir(PATH_FOLDER_DOWNLOAD):
                ctime = os.path.getctime(os.path.join(PATH_FOLDER_DOWNLOAD, name_file_download))
                list_name_ctime.append((name_file_download, ctime))
            recent_file = max(list_name_ctime, key=lambda x : x[1])[0]
            X = pd.DataFrame(pd.read_excel(os.path.join(PATH_FOLDER_DOWNLOAD, recent_file), sheet_name='운행내역(콜)'))
            X_unique = X[~X.apply(tuple, axis=1).reset_index(drop=True).isin(df_rawdata.apply(tuple, axis=1).reset_index(drop=True))]
            df_rawdata = pd.concat([df_rawdata, X_unique], ignore_index=True)
            df_rawdata = pd.concat([df_rawdata, X]).drop_duplicates(subset=df_rawdata.columns[:5].tolist() + df_rawdata.columns[6:21].tolist(), keep='last')
            os.remove(os.path.join(PATH_FOLDER_DOWNLOAD, recent_file))
        df_rawdata.to_csv(PATH_RAWDATA, index=False, encoding='utf-8')
        driver.quit()
        def rawdata_to_for_excel():

            import pandas as pd
            A = pd.read_csv(PATH_RAWDATA, encoding='utf-8')
            B = pd.read_csv(PATH_HOLIDAY, encoding='euc-kr')

            import math
            def round_down_to_30_min(time_str):
                if pd.isna(time_str):  # NaN 처리
                    return "00:00:00"
                if isinstance(time_str, float):  # float일 경우 문자열로 변환
                    time_str = str(time_str)
                h, m, s = map(int, time_str.split(":"))
                rounded_minutes = math.floor(m / 30) * 30
                return f"{h:02}:{rounded_minutes:02}:00"
            def convert_to_minute(time_str):
                if pd.isna(time_str):  # NaN 처리
                    return "-"
                if isinstance(time_str, float):  # float일 경우 문자열로 변환
                    time_str = str(time_str)
                h, m, s = map(int, time_str.split(":"))
                converted_minute = round(h*60 + m + s/60, 1)
                return f"{converted_minute}"

            import datetime
            A['날짜'] = pd.to_datetime(A['날짜'], format='%Y-%m-%d')
            B['날짜'] = pd.to_datetime(B['날짜'])
            cut_off_date = datetime.datetime(2024, 7, 15)
            A['사업 구분'] = A['날짜'].apply(lambda x: '시범사업' if x < cut_off_date else '본사업')
            A['주중/주말'] = A['날짜'].map(B.set_index('날짜')['주중/주말'])
            A['이용인원'] = A['성인'] + A['청소년'] + A['어린이']
            A['호출시각'] = A['호출시각'].apply(round_down_to_30_min)
            A['차량 도착 시각 (픽업)'] = A['차량 도착 시각 (픽업)'].apply(round_down_to_30_min)
            A['대기 시간\n(차량 도착 시각-호출 시각)'] = A['대기 시간\n(차량 도착 시각-호출 시각)'].apply(convert_to_minute)

            A[['지역(지자체)','이용인원','배차분류','호출방법','호출시각','차량 도착 시각 (픽업)','대기 시간\n(차량 도착 시각-호출 시각)','날짜', '사업 구분', '주중/주말']].to_csv(PATH_DB_FOR_EXCEL, index=False, encoding='euc-kr')
        # df_rawdata[['지역(지자체)','성인','청소년','어린이','배차분류','날짜','호출방법','호출시각','차량 도착 시각 (픽업)','하차시각','대기 시간\n(차량 도착 시각-호출 시각)']].to_csv(PATH_DB_FOR_EXCEL, index=False, encoding= 'utf-8')
        rawdata_to_for_excel()

        import xlwings as xw
        app = xw.App(visible=True)
        wb = app.books.open(PATH_EXCEL)
        sheet = wb.sheets['제주옵서버스DB']
        from datetime import datetime, timedelta
        from_date = datetime.strptime(FROM_DATE, "%Y-%m-%d")
        to_date = datetime.strptime(TO_DATE, "%Y-%m-%d")
        plus_bag = from_date
        dates = []
        while plus_bag <= from_date + timedelta(days=6):
            dates.append(plus_bag.strftime("%Y-%m-%d"))
            plus_bag += timedelta(days=1)

        for idx, i in enumerate(dates):
            num_row = idx + 29
            sheet[f'N{num_row}'].value = i

    btn = Button(win, width=15, height=1, command=update_DB ,font=Font(family = "KoPub돋움체 bold", size=15), text=label_text)
    btn.grid(column=col, row=row, padx=0, pady=10, columnspan=2, rowspan=2)

btn_DB업데이트 = create_button_and_entry_update_DB(win, "DB업데이트", row=11, col=0)

def create_button_and_entry_fill_report(win, label_text, row, col):
    def fill_report():
        PATH_FILE_REPORT = ent_한글_보고서.get()
        PATH_EXCEL = ent_Excel_계산용.get()

        def put_data_to_report(num_pos_row, num_pos_col, len_row, len_col, name_field, how_round):
            for row in range(0, len_row):
                for col in range(0, len_col):
                    if name_field == '일일_이용현황' or name_field == '본사업_누적_이용현황_값':
                        num_continue_row = (row + 6) % 7
                        if num_continue_row == 0 and col == 2:
                            continue
                        import math # row 1, 8, 15, 22, 29, 36, 43 / col 3 을 pass 하도록
                        num_field = row*len_col + col - math.ceil((row*len_col + col - 9)/49) 
                    else:
                        num_field = row*len_col + col
                    if isinstance(ws.cell(row = row+num_pos_row, column = col+num_pos_col).value, float) == True or isinstance(ws.cell(row = row+num_pos_row, column = col+num_pos_col).value, int) == True:
                        print(f'{name_field}{{{{{num_field}}}}}')
                        print(round(ws.cell(row = row+num_pos_row, column = col+num_pos_col).value, how_round))
                        if name_field == '본사업_누적_이용현황_비율' or name_field == '일자별_콜인입_현황_총응답률':
                            hwp.put_field_text(f'{name_field}{{{{{num_field}}}}}',f'{round(ws.cell(row = row+num_pos_row, column = col+num_pos_col).value*100, how_round)}%')
                        else:
                            if name_field == '일일_대기시간_현황' or name_field == '주중_시간대별_일평균_수요_현황':
                                hwp.put_field_text(f'{name_field}{{{{{num_field}}}}}',round(ws.cell(row = row+num_pos_row, column = col+num_pos_col).value, how_round))
                            else:
                                hwp.put_field_text(f'{name_field}{{{{{num_field}}}}}',f'{ws.cell(row = row+num_pos_row, column = col+num_pos_col).value:,}')
                                

                    else:
                        print(f'{name_field}{{{{{num_field}}}}}')
                        print(ws.cell(row = row+num_pos_row, column = col+num_pos_col).value)
                        hwp.put_field_text(f'{name_field}{{{{{num_field}}}}}', ws.cell(row = row+num_pos_row, column = col+num_pos_col).value)
        def add_row_field(add_row, name_field):
            import pyautogui as pag
            import time
            # 선택 블록 이동 _ 합계 윗 칸
            hwp.set_pos(1040, 0, 0)
            pag.keyDown('alt')
            time.sleep(0.5)
            pag.press('c')
            time.sleep(0.5)
            pag.press('b')
            time.sleep(0.5)
            pag.press('d')
            time.sleep(0.5)
            pag.keyUp('alt')
            time.sleep(0.5)
            hwp.move_to_field('일자별_이용완료_인원/건수_현황_맨앞')
            pag.hotkey('f5')
            pag.hotkey('f5')
            pag.hotkey('f5')
            pag.hotkey('f5')
            pag.hotkey('home')
            pag.press('up')
            # 7 줄 추가
            pag.keyDown('alt')
            pag.hotkey('insert')
            time.sleep(0.5)
            pag.press('b')
            pag.press('c')
            pag.keyUp('alt')
            pag.write(add_row)
            pag.hotkey('enter')
            time.sleep(0.5)
            # 필드지정 _ 
            pag.press('down')
            pag.hotkey('f5')
            pag.hotkey('end')
            pag.hotkey('pgdn')
            pag.press('p')
            time.sleep(0.5)
            pag.hotkey('end')
            pag.keyDown('alt')
            pag.press('n')
            pag.keyUp('alt')
            import pyperclip
            pyperclip.copy(name_field) #'일자별_이용완료_인원/건수_현황'
            pag.keyDown('ctrl')
            pag.press('v')
            pag.keyUp('ctrl')
            pag.hotkey('enter')
        from hwpWings import HWP

        hwp = HWP()
        hwp.open(PATH_FILE_REPORT)
        from openpyxl import load_workbook
        wb = load_workbook(PATH_EXCEL, data_only=True)
        ws = wb['제주옵서버스DB']
        hwp.set_pos(0, 1, 1)
        import pyautogui as pag
        for i in range(0, 11):
            pag.press("del")
        hwp.insert_text(datetime.today().strftime("%Y.%m.%d."))

        hwp.set_pos(0, 7, 13)
        for i in range(0, 11):
            pag.press("del")
        hwp.insert_text(end_date.strftime("%Y.%m.%d."))
        
        add_row_field('7', '일자별_이용완료_인원/건수_현황')
        put_data_to_report(129, 16, 8, 13, '일자별_이용완료_인원/건수_현황', 0)
        put_data_to_report(11, 17, 7, 7,'본사업_누적_이용현황_값',0)
        put_data_to_report(18, 18, 7, 5,'본사업_누적_이용현황_비율',1)
        put_data_to_report(29, 17, 49, 7, '일일_이용현황', 0)
        put_data_to_report(37, 14, 17, 1, '한림읍_취소건수', 0)
        put_data_to_report(81, 17, 42, 3, '일일_대기시간_현황', 1)
        put_data_to_report(4, 27, 8, 3,'일자별_콜인입_현황_값', 0)
        put_data_to_report(4, 30, 8, 1,'일자별_콜인입_현황_총응답률', 1)
        put_data_to_report(4, 31, 8, 8, '일일_주요문의건', 0)
        put_data_to_report(141, 17, 16, 6, '주중_시간대별_일평균_수요_현황', 1)
        put_data_to_report(29, 13, 7, 1, '날짜', 0)
        put_data_to_report(29, 12, 7, 1, '날짜1', 0)
        put_data_to_report(29, 13, 7, 1, '날짜2', 0)
        put_data_to_report(29, 13, 7, 1, '날짜3', 0)

        def Textcolor_change():
            try:
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableUpperCell")
                hwp.HAction.Run("TableUpperCell")
                hwp.HAction.Run("TableUpperCell")
                hwp.HAction.Run("TableCellBlockExtend")
                hwp.HAction.Run("TableColBegin")
                hwp.HAction.Run("TableUpperCell")

                hwp.HAction.GetDefault("ShapeCopyPaste", hwp.HParameterSet.HShapeCopyPaste.HSet)
                hwp.HAction.Execute("ShapeCopyPaste", hwp.HParameterSet.HShapeCopyPaste.HSet)
            except Exception as e:
                print(f"오류 발생: {e}")

        import time
        time.sleep(0.5)
        pag.press('p')
        pag.keyDown('alt')
        pag.press('n')
        pag.keyUp('alt')
        pag.hotkey('del')
        pag.hotkey('enter')

        Textcolor_change()

        time.sleep(1)
        pag.keyDown('alt')
        pag.press('c')
        pag.keyUp('alt')
        time.sleep(1)

    btn = Button(win, width=15, height=1, command=fill_report ,font=Font(family = "KoPub돋움체 bold", size=15), text=label_text)
    btn.grid(column=col, row=row, padx=0, pady=10, columnspan=2, rowspan=2)

btn_보고서작성 = create_button_and_entry_fill_report(win, "보고서작성", row=11, col=2)

sheet_data = []

def update_combobox():
    combo_box['values'] = sheet_data

def create_button_and_entry_get_gid(win, label_text, row, col):
    def get_gid():
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.chrome.service import Service
        from webdriver_manager.chrome import ChromeDriverManager
        import time
        import re
        from tkinter import Button, messagebox

        # ChromeDriver 설정
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))

        # URL로 이동 (예시 URL)
        driver.get("https://docs.google.com/spreadsheets/d/1HuYVykidqTv30vigjr7tLBOU6vOeqri8/edit?gid=")

        # 페이지 로드 후 처음 URL 및 이름 추출
        if "제주" in driver.current_url:
            # 처음 로드된 URL 추출
            current_url = driver.current_url

            # gid 값 추출
            initial_match = re.search(r"gid=(\d+)", current_url)
            if initial_match:
                gid_value = initial_match.group(1)
            else:
                gid_value = "Unknown"

            # 처음 로드된 시트 이름 추출 (활성화된 탭 탐색)
            active_tab = driver.find_element(By.XPATH, "//div[@class='goog-inline-block docs-sheet-tab docs-material docs-sheet-active-tab']")
            element_text = active_tab.text  # 텍스트(시트 이름) 추출
            
            sheet_data.append((element_text, gid_value))


        # XPATH에 해당하는 요소들 찾기 (다른 시트 탭 정보)
        elements = driver.find_elements(By.XPATH, "//div[@class='goog-inline-block docs-sheet-tab docs-material']")

        # 각 요소에서 id와 텍스트 추출
        for element in elements:
            element_id = element.get_attribute('id')  # id 추출
            element_text = element.text.strip()  # 텍스트 추출

            # 텍스트가 비어 있지 않으면 출력
            if "제주" in element_text and "문의" not in element_text:
                try:
                    # 시트 클릭
                    sheet_tab = driver.find_element(By.ID, element_id)
                    sheet_tab.click()

                    # 클릭 후 URL 추출
                    current_url = driver.current_url

                    # URL에서 gid 값 추출
                    match = re.search(r"gid=(\d+)", current_url)
                    if match:
                        gid_value = match.group(1)
                        sheet_data.append((element_text, gid_value))
                        print(f"{element_text} : {gid_value}")

                except Exception as e:
                    print(f"시트를 클릭하는 데 오류가 발생했습니다: {e}")

        # 브라우저 종료
        driver.quit()
        update_combobox()
        messagebox.showinfo("알림", "GID 추출이이 완료되었습니다.")

    btn = Button(win, command=get_gid, width=12, height=1, anchor="w", font=Font(family="KoPub돋움체 bold", size = 10), text=label_text)
    btn.grid(column=col, row=row, padx=20, pady=5)

btn_get_gid = create_button_and_entry_get_gid(win, "GID 추출", row=10, col=0)

import tkinter as tk
combo_var = tk.StringVar
combo_box = ttk.Combobox(win, textvariable=combo_var, width=18)
combo_box.grid(row = 10, column= 1, columnspan=2, pady=10)

def create_button_and_entry_adopt_query_gid(win, label_text, row, col):
    def adopt_query_gid():
        import xlwings as xw
        import re
        from tkinter import Button, messagebox

        # 엑셀 파일 경로
        file_path = ent_Excel_계산용.get()

        # 엑셀 열기
        wb = xw.Book(file_path)

        # Power Query 가져오기
        queries = wb.api.Queries

        # 모든 sheet_data를 순회하면서 처리
        for element_text, gid_value in sheet_data:
            # element_text에서 월을 숫자로 추출 (정규식을 사용하여 "2월", "3월" 형식으로 추출)
            month_match = re.search(r'(\d+)월', element_text)
            if month_match:
                month = int(month_match.group(1))  # 월을 숫자로 변환
            else:
                print(f"'{element_text}'에서 월을 추출할 수 없습니다.")
                continue  # 월 추출 실패 시 다음으로 넘어감

            # 홀수/짝수 월에 맞는 쿼리 이름들
            if month % 2 == 0:  # 짝수월일 경우
                for query in queries:
                    query_name = query.Name
                    if query_name == "짝수달_콜현황보고(제주)_콜인입" or query_name == "짝수달_콜현황보고(제주)_주요문의":
                        # gid=(\d+) 부분을 gid_value로 대체
                        modified_formula = re.sub(r"gid=(\d+)", f"gid={gid_value}", query.Formula)
                        # 수정된 쿼리 공식 적용
                        query.Formula = modified_formula  # 수정된 쿼리 공식 반영
                        print(f"변경된 쿼리: {modified_formula}")
                        
            else:  # 홀수월일 경우
                for query in queries:
                    query_name = query.Name
                    if query_name == "홀수달_콜현황보고(제주)_콜인입" or query_name == "홀수달_콜현황보고(제주)_주요문의":
                        # gid=(\d+) 부분을 gid_value로 대체
                        modified_formula = re.sub(r"gid=(\d+)", f"gid={gid_value}", query.Formula)
                        # 수정된 쿼리 공식 적용
                        query.Formula = modified_formula  # 수정된 쿼리 공식 반영
                        print(f"변경된 쿼리: {modified_formula}")

        # 엑셀 파일 저장 (변경된 내용 반영)
        wb.save()
        print("엑셀 파일에 변경된 쿼리가 적용되었습니다.")
        messagebox.showinfo("알림", "GID 업데이트가 완료되었습니다.")

    btn = Button(win, command=adopt_query_gid, width=12, height=1, anchor="w", font=Font(family="KoPub돋움체 bold", size = 10), text=label_text)
    btn.grid(column=col, row=row, padx=20, pady=5)

btn_get_gid = create_button_and_entry_adopt_query_gid(win, "GID 업데이트", row=10, col=3)

import tkinter as tk
from tkinter import font, messagebox
from tkcalendar import Calendar
import csv

# 전역 변수
holiday_date = []

# 주중/주말 데이터를 읽는 함수
def load_weekend_weekday_data():
    weekdays = {}
    with open(ent_DB_주중주말.get(), mode="r", encoding="euc-kr") as file:
        reader = csv.reader(file)
        for row in reader:
            date = row[0]
            weekday_or_weekend = row[1]
            weekdays[date] = weekday_or_weekend
    return weekdays

# 날짜 추가 함수
def add_date_to_list():
    selected_date = cal.get_date()
    # 주중/주말 정보를 불러옴
    weekdays = load_weekend_weekday_data()
    weekday_or_weekend = weekdays.get(selected_date, "주중")  # 기본값은 "주중"
    
    if selected_date not in holiday_date:
        holiday_date.append(selected_date)
        # 날짜와 주중/주말 정보를 텍스트 박스에 추가
        date_display.insert(tk.END, f"{selected_date} ({weekday_or_weekend})\n")
        
# 날짜 삭제 함수
def remove_date_from_list(date):
    if date in holiday_date:
        holiday_date.remove(date)
        display_dates()

# 전체 삭제 함수
def clear_all_dates():
    holiday_date.clear()
    display_dates()

# "변경 완료" 버튼 클릭 시 주중/주말 값 변경 함수
def change_to_weekend():
    weekdays = load_weekend_weekday_data()
    updated_rows = []
    
    # CSV 파일을 읽어들여 모든 데이터를 저장
    with open(ent_DB_주중주말.get(), mode="r", encoding="euc-kr") as file:
        reader = csv.reader(file)
        updated_rows = list(reader)  # 데이터를 리스트로 읽음
    
    # holiday_date에 있는 날짜들을 "주말"로 변경
    for i, row in enumerate(updated_rows):
        if row[0] in holiday_date:
            updated_rows[i][1] = "주말"  # 해당 날짜의 우측 데이터를 "주말"로 변경
    
    # 변경된 내용을 다시 CSV 파일에 저장
    with open(ent_DB_주중주말.get(), mode="w", encoding="euc-kr", newline="") as file:
        writer = csv.writer(file)
        writer.writerows(updated_rows)  # 변경된 내용 저장

# 날짜 리스트를 화면에 표시하는 함수
def display_dates(message=None):
    date_display.delete(1.0, tk.END)  # 기존 내용 삭제
    if message:
        date_display.insert(tk.END, f"{message}\n")  # 변경 완료 메시지 출력
    for date in holiday_date:
        # 주중/주말 정보를 텍스트 박스에 출력
        weekdays = load_weekend_weekday_data()
        weekday_or_weekend = weekdays.get(date, "주중")  # 기본값은 "주중"
        date_display.insert(tk.END, f"{date} ({weekday_or_weekend})\n")

# 달력 창 생성 함수
def open_calendar_window():
    # 이미 열린 달력 창이 있으면 새로 열지 않도록 체크
    if hasattr(open_calendar_window, "calendar_window") and open_calendar_window.calendar_window.winfo_exists():
        return

    open_calendar_window.calendar_window = tk.Toplevel(win)
    open_calendar_window.calendar_window.title("날짜 선택")

    # 달력 생성
    global cal
    cal = Calendar(open_calendar_window.calendar_window, selectmode="day", date_pattern="yyyy-mm-dd")
    cal.grid(row=0, column=0, padx=20, pady=5, columnspan=3)  # 달력의 너비를 3개의 열로 확장

    # 날짜 추가 후 추가 버튼
    add_button = tk.Button(open_calendar_window.calendar_window, text="추가", command=add_date_to_list)
    add_button.grid(row=1, column=0, padx=20, pady=5, sticky="ew", columnspan=1)  # 버튼을 3개의 열로 확장

    # 삭제 버튼
    cancel_button = tk.Button(open_calendar_window.calendar_window, text="취소", command=lambda: remove_date_from_list(cal.get_date()))
    cancel_button.grid(row=1, column=1, padx=5, pady=5, sticky="ew", columnspan=1)  # 버튼을 3개의 열로 확장

    # 전체 삭제 버튼
    clear_all_button = tk.Button(open_calendar_window.calendar_window, text="전체 취소", command=clear_all_dates)
    clear_all_button.grid(row=1, column=2, padx=20, pady=5, sticky="ew", columnspan=1)  # 버튼을 3개의 열로 확장

    # 선택된 날짜가 추가된 텍스트 박스
    global date_display
    date_display = tk.Text(open_calendar_window.calendar_window, height=10, width=30)
    date_display.grid(row=2, column=0, padx=20, pady=5, columnspan=3)

    # 변경 완료 버튼
    change_button = tk.Button(open_calendar_window.calendar_window, text="주중 → 주말 변경하기", command=change_to_weekend, width=30)
    change_button.grid(row=3, column=0, padx=20, pady=5, columnspan=3)

# 주중주말선택 버튼 생성
btn = tk.Button(win, text="주말 추가하기", command=open_calendar_window, width=12, height=1, anchor="w", font=font.Font(family="KoPub돋움체 bold", size=10))
btn.grid(column=3, row=8, padx=20, pady=10)
 
win.mainloop()




